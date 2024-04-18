from openpyxl import load_workbook

# Excel 파일 불러오기
workbook = load_workbook(filename="c:/work3/products.xlsx")
sheet = workbook.active

# 제품 목록 읽어오기
products = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    product_id, product_name, quantity, price = row
    products.append({
        "제품ID": product_id,
        "제품명": product_name,
        "수량": quantity,
        "가격": price
    })

# 제품 목록 출력
for product in products:
    print(product)
